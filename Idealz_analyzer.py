"""
╔══════════════════════════════════════════════════════════════════════╗
║       IDEALZ GOOGLE REVIEWS — ANALYZER v2 (IMPROVED)                ║
╚══════════════════════════════════════════════════════════════════════╝
RUN:   python idealz_analyzer_v2.py
OUTPUT: idealz_reviews_report.xlsx
"""

import json, re
from pathlib import Path
from datetime import datetime
from collections import Counter

import pandas as pd
from textblob import TextBlob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

INPUT_JSON  = "idealz_raw_reviews.json"
OUTPUT_XLSX = "idealz_reviews_report.xlsx"

# ── CATEGORY KEYWORDS ─────────────────────────────────────────────────────────
CATEGORIES = {
    "Product Quality":    ["quality", "product", "item", "defect", "broken", "material",
                           "durable", "genuine", "fake", "original", "condition", "damaged"],
    "Customer Service":   ["service", "staff", "employee", "helpful", "rude", "assist",
                           "support", "polite", "attitude", "friendly", "ignored",
                           "customer care", "respond", "behavior", "behaviour"],
    "Pricing & Value":    ["price", "pricing", "expensive", "cheap", "affordable", "value",
                           "worth", "overpriced", "discount", "offer", "deal", "cost", "money"],
    "Delivery":           ["delivery", "shipping", "courier", "late", "arrived", "dispatch",
                           "tracking", "fast delivery", "delayed", "on time", "deliver"],
    "Store Experience":   ["store", "shop", "location", "parking", "clean", "organized",
                           "display", "ambiance", "atmosphere", "interior", "mall",
                           "crowded", "visit", "branch", "outlet"],
    "Online / App":       ["website", "app", "online", "order", "checkout", "payment",
                           "login", "cart", "interface", "loading", "web", "platform"],
    "Promotions / Deals": ["promotion", "sale", "offer", "discount", "coupon", "voucher",
                           "cashback", "deal", "lucky draw", "win", "prize", "raffle"],
    "After-Sales":        ["return", "refund", "exchange", "warranty", "replace",
                           "complaint", "issue", "resolve", "compensation", "claim"],
    "Variety / Stock":    ["variety", "selection", "stock", "available", "range",
                           "choice", "collection", "assortment", "limited", "out of stock"],
}

# ── COLOUR PALETTE ────────────────────────────────────────────────────────────
C = {
    "navy":    "1F3864", "blue":    "2E75B6", "dkblue":  "1A4F8A",
    "lblue":   "D6E4F0", "green":   "1E5631", "lgreen":  "C6EFCE",
    "red":     "9C0006", "lred":    "FFC7CE", "lyellow": "FFEB9C",
    "orange":  "ED7D31", "lgray":   "F2F2F2", "altrow":  "EBF3FB",
    "white":   "FFFFFF", "border":  "BFBFBF",
}
SENT_CLR  = {"Positive": C["lgreen"], "Negative": C["lred"],
             "Neutral":  C["lyellow"], "No Text":  C["lgray"]}
STAR_CLR  = {5: "C6EFCE", 4: "DDEBF7", 3: "FFF2CC", 2: "FCE4D6", 1: "FFC7CE"}
STORE_CLR = ["1F4E79", "375623", "7B2C2C"]   # navy / dark green / dark red per store

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
def _border():
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, r1, r2, c1, c2):
    b = _border()
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cell in row:
            cell.border = b

def hdr(cell, text, bg=None, fg="FFFFFF", sz=11, bold=True, center=True):
    cell.value = text
    cell.font  = Font(name="Arial", bold=bold, size=sz, color=fg)
    cell.fill  = PatternFill("solid", start_color=bg or C["navy"])
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=True)

def dat(cell, value, center=False, bold=False, sz=10, fg="000000", indent=0):
    cell.value = value
    cell.font  = Font(name="Arial", size=sz, bold=bold, color=fg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", indent=indent, wrap_text=True)

def alt(ws, row, ncols, c1=1):
    if row % 2 == 0:
        for c in range(c1, c1 + ncols):
            ws.cell(row, c).fill = PatternFill("solid", start_color=C["altrow"])

def set_col_widths(ws, widths):
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

def banner(ws, text, merge_to="J1", bg=None, sz=15, row=1):
    ws.merge_cells(f"A{row}:{merge_to}")
    hdr(ws[f"A{row}"], text, bg=bg or C["navy"], sz=sz)
    ws.row_dimensions[row].height = 42

def sub_banner(ws, text, merge_to, row, bg=None):
    ws.merge_cells(f"A{row}:{merge_to}{row}")
    ws[f"A{row}"].value = text
    ws[f"A{row}"].font  = Font(name="Arial", size=9, italic=True, color="555555")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 18

# ── ANALYSIS ──────────────────────────────────────────────────────────────────

def parse_date(raw: str) -> str:
    """
    Convert Google's relative timestamps to YYYY-MM.
    Handles: 'X days/weeks/months/years ago', 'a month ago', 'a year ago',
             'yesterday', 'just now', 'today', 'an hour ago'
    """
    if not raw:
        return ""
    r = raw.lower().strip()
    today = datetime.today()

    # "just now" / "today" / "an hour ago" / "a few seconds ago"
    if any(x in r for x in ["just now", "today", "hour ago", "second", "minute"]):
        return today.strftime("%Y-%m")
    if "yesterday" in r:
        return (today - pd.DateOffset(days=1)).strftime("%Y-%m")

    # "a week ago" / "a month ago" / "a year ago"
    singular = re.match(r"^a\s+(day|week|month|year)\s+ago$", r)
    if singular:
        unit = singular.group(1)
        offset = {"day": pd.DateOffset(days=1), "week": pd.DateOffset(weeks=1),
                  "month": pd.DateOffset(months=1), "year": pd.DateOffset(years=1)}
        return (today - offset[unit]).strftime("%Y-%m")

    # "3 months ago" / "2 years ago" etc.
    m = re.search(r"(\d+)\s+(day|week|month|year)", r)
    if m:
        n, unit = int(m.group(1)), m.group(2)
        offset = {"day": pd.DateOffset(days=n), "week": pd.DateOffset(weeks=n),
                  "month": pd.DateOffset(months=n), "year": pd.DateOffset(years=n)}
        return (today - offset[unit]).strftime("%Y-%m")

    return ""

def categorize(text: str) -> str:
    if not text or not isinstance(text, str):
        return "General"
    tl = text.lower()
    matched = [cat for cat, kws in CATEGORIES.items() if any(kw in tl for kw in kws)]
    return ", ".join(matched) if matched else "General"

def get_sentiment(text: str) -> tuple:
    if not text or not isinstance(text, str) or not text.strip():
        return ("No Text", 0.0)
    score = TextBlob(text).sentiment.polarity
    label = "Positive" if score > 0.1 else ("Negative" if score < -0.1 else "Neutral")
    return (label, round(score, 3))

def build_df(raw: list) -> pd.DataFrame:
    rows = []
    for r in raw:
        sl, ss = get_sentiment(r.get("text", ""))
        my = parse_date(r.get("date_raw", ""))
        rows.append({
            "store":    r.get("store", ""),
            "author":   r.get("author", "Anonymous"),
            "rating":   r.get("rating"),
            "date_raw": r.get("date_raw", ""),
            "month_year": my,
            "year":     my[:4] if len(my) == 7 else "",
            "text":     r.get("text", ""),
            "likes":    r.get("likes", 0),
            "owner_reply": r.get("owner_reply", ""),
            "has_reply":   "Yes" if str(r.get("owner_reply","")).strip() else "No",
            "sentiment":   sl,
            "sent_score":  ss,
            "categories":  categorize(r.get("text", "")),
        })
    df = pd.DataFrame(rows)
    df["rating"] = pd.to_numeric(df["rating"], errors="coerce")
    return df

# ── SHEET 1: DASHBOARD ────────────────────────────────────────────────────────

def write_dashboard(wb, df):
    ws = wb.create_sheet("📊 Dashboard")
    ws.sheet_view.showGridLines = False

    banner(ws, "IDEALZ STORES — GOOGLE REVIEWS ANALYSIS DASHBOARD",
           merge_to="L1", sz=16)
    sub_banner(ws, (f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}   |   "
                    f"Total Reviews: {len(df)}   |   Stores: {df['store'].nunique()}   |   "
                    f"Data as of: 22 February 2026"),
               merge_to="L", row=2)

    stores = list(df["store"].unique())

    # ── KPI cards — 4 columns per store ──
    ROW = 4
    for si, sname in enumerate(stores):
        sdf  = df[df["store"] == sname]
        avg  = sdf["rating"].mean()
        pos  = 100 * (sdf["sentiment"] == "Positive").sum() / max(len(sdf), 1)
        neg  = 100 * (sdf["sentiment"] == "Negative").sum() / max(len(sdf), 1)
        rep  = 100 * (sdf["has_reply"] == "Yes").sum() / max(len(sdf), 1)
        sc   = STORE_CLR[si % len(STORE_CLR)]
        c1   = 1 + si * 4

        ws.merge_cells(start_row=ROW, start_column=c1, end_row=ROW, end_column=c1+3)
        hdr(ws.cell(ROW, c1), sname, bg=sc, sz=10)
        ws.row_dimensions[ROW].height = 26

        kpis = [
            ("Total Reviews",   f"{len(sdf):,}",          sc),
            ("Avg Rating",      f"{avg:.2f} ★" if pd.notna(avg) else "N/A",  "B8860B"),
            ("5★ Reviews",      f"{int((sdf['rating']==5).sum()):,}",         "1E5631"),
            ("% Positive",      f"{pos:.0f}%",             "1E5631"),
            ("% Negative",      f"{neg:.0f}%",             "9C0006"),
            ("Owner Replies",   f"{rep:.0f}%",             "444444"),
        ]
        for ki, (label, val, color) in enumerate(kpis):
            r = ROW + 1 + ki
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c1+2)
            lc = ws.cell(r, c1)
            lc.value = label
            lc.font  = Font(name="Arial", size=9, color="444444")
            lc.fill  = PatternFill("solid", start_color="F5F8FC")
            lc.alignment = Alignment(horizontal="left", indent=1, vertical="center")

            vc = ws.cell(r, c1+3)
            vc.value = val
            vc.font  = Font(name="Arial", size=12, bold=True, color=color)
            vc.fill  = PatternFill("solid", start_color="F5F8FC")
            vc.alignment = Alignment(horizontal="center", vertical="center")
            apply_border(ws, r, r, c1, c1+3)

    # ── Store comparison table ──
    ROW = 13
    ws.merge_cells(f"A{ROW}:L{ROW}")
    hdr(ws.cell(ROW, 1), "STORE COMPARISON", bg=C["dkblue"], sz=11)
    ws.row_dimensions[ROW].height = 26
    ROW += 1

    hdrs = ["Store", "Total", "Avg ★", "5★", "4★", "3★", "2★", "1★",
            "Positive %", "Negative %", "Neutral %", "Owner Reply %"]
    for ci, h in enumerate(hdrs, 1):
        hdr(ws.cell(ROW, ci), h, bg="365F91", sz=10)
    ws.row_dimensions[ROW].height = 24
    tbl_s = ROW
    ROW += 1

    for si, sname in enumerate(stores):
        sdf = df[df["store"] == sname]
        rc  = sdf["rating"].value_counts()
        pos = 100*(sdf["sentiment"]=="Positive").sum()/max(len(sdf),1)
        neg = 100*(sdf["sentiment"]=="Negative").sum()/max(len(sdf),1)
        neu = 100*(sdf["sentiment"]=="Neutral").sum()/max(len(sdf),1)
        rep = 100*(sdf["has_reply"]=="Yes").sum()/max(len(sdf),1)
        avg = sdf["rating"].mean()
        vals = [sname, len(sdf),
                f"{avg:.2f}" if pd.notna(avg) else "N/A",
                rc.get(5,0), rc.get(4,0), rc.get(3,0), rc.get(2,0), rc.get(1,0),
                f"{pos:.1f}%", f"{neg:.1f}%", f"{neu:.1f}%", f"{rep:.1f}%"]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ROW, ci)
            dat(cell, v, center=(ci > 1))
            if ROW % 2 == 0:
                cell.fill = PatternFill("solid", start_color=C["altrow"])
        ROW += 1

    apply_border(ws, tbl_s, ROW-1, 1, 12)

    # Column widths
    set_col_widths(ws, [28, 8, 8, 7, 7, 7, 7, 7, 11, 11, 10, 13])


# ── SHEET 2: ALL REVIEWS ──────────────────────────────────────────────────────

def write_raw(wb, df):
    ws = wb.create_sheet("📋 All Reviews")
    ws.sheet_view.showGridLines = False

    hdrs   = ["Store","Author","Rating","Date (Raw)","Month-Year","Sentiment",
              "Sent. Score","Categories","Review Text","Likes","Owner Reply?","Owner Reply"]
    fields = ["store","author","rating","date_raw","month_year","sentiment",
              "sent_score","categories","text","likes","has_reply","owner_reply"]
    widths = [24, 18, 8, 18, 12, 12, 10, 35, 65, 7, 13, 50]

    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        hdr(ws.cell(1, ci), h, bg=C["navy"], sz=10)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26

    for ri, (_, row) in enumerate(df[fields].iterrows(), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            cell.value = val
            cell.font  = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == 9))

        sent = str(row["sentiment"])
        if sent in SENT_CLR:
            ws.cell(ri, 6).fill = PatternFill("solid", start_color=SENT_CLR[sent])

        rat = row["rating"]
        if pd.notna(rat) and int(rat) in STAR_CLR:
            ws.cell(ri, 3).fill = PatternFill("solid", start_color=STAR_CLR[int(rat)])
            ws.cell(ri, 3).alignment = Alignment(horizontal="center", vertical="top")

        alt(ws, ri, len(fields))

    apply_border(ws, 1, len(df)+1, 1, len(fields))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}1"


# ── SHEET 3: RATINGS ──────────────────────────────────────────────────────────

def write_ratings(wb, df):
    ws = wb.create_sheet("⭐ Ratings")
    ws.sheet_view.showGridLines = False
    banner(ws, "RATINGS DISTRIBUTION", merge_to="H1", sz=14)
    sub_banner(ws, f"Based on {int(df['rating'].count())} reviews with star ratings",
               "H", row=2)

    stores = list(df["store"].unique())
    all_label = "ALL STORES"
    cols = stores + [all_label]

    ROW = 4
    hdr(ws.cell(ROW, 1), "Stars", bg=C["blue"], sz=10)
    for ci, s in enumerate(cols, 2):
        hdr(ws.cell(ROW, ci), s, bg=C["blue"] if s != all_label else C["navy"], sz=10)
    ws.row_dimensions[ROW].height = 24
    data_row_start = ROW + 1
    ROW += 1

    for stars in [5, 4, 3, 2, 1]:
        cell = ws.cell(ROW, 1)
        cell.value = "★" * stars
        cell.font  = Font(name="Arial", size=12, bold=True,
                          color=("FFB700" if stars >= 4 else ("FF4444" if stars <= 2 else "AA8800")))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", start_color=STAR_CLR[stars])

        for ci, s in enumerate(cols, 2):
            sub  = df if s == all_label else df[df["store"] == s]
            cnt  = int((sub["rating"] == stars).sum())
            tot  = sub["rating"].count()
            pct  = cnt / max(tot, 1) * 100
            cell = ws.cell(ROW, ci)
            cell.value = f"{cnt:,}   ({pct:.1f}%)"
            cell.font  = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill  = PatternFill("solid", start_color=STAR_CLR[stars])
        ROW += 1

    # Totals row
    hdr(ws.cell(ROW, 1), "TOTAL", bg="365F91", sz=10)
    for ci, s in enumerate(cols, 2):
        sub = df if s == all_label else df[df["store"] == s]
        hdr(ws.cell(ROW, ci), f"{int(sub['rating'].count()):,}", bg="365F91", sz=10)
    ROW += 1
    apply_border(ws, data_row_start - 1, ROW - 1, 1, len(cols) + 1)

    set_col_widths(ws, [10] + [22] * len(cols))

    # ── Bar chart — raw counts ──
    ROW += 2
    chart = BarChart()
    chart.type  = "col"
    chart.title = "Star Rating Distribution by Store"
    chart.y_axis.title = "Number of Reviews"
    chart.style = 10; chart.width = 28; chart.height = 16

    # Write numeric-only data for chart (separate from formatted cells)
    chart_row = ROW
    hdr(ws.cell(chart_row, 10), "Stars", bg=C["navy"], sz=9)
    for ci, s in enumerate(cols, 11):
        hdr(ws.cell(chart_row, ci), s, bg=C["navy"], sz=9)
    chart_row += 1
    chart_data_start = chart_row
    for stars in [5, 4, 3, 2, 1]:
        ws.cell(chart_row, 10).value = f"{stars}★"
        for ci, s in enumerate(cols, 11):
            sub = df if s == all_label else df[df["store"] == s]
            ws.cell(chart_row, ci).value = int((sub["rating"] == stars).sum())
        chart_row += 1

    data_ref = Reference(ws, min_col=11, max_col=10+len(cols),
                         min_row=chart_data_start-1, max_row=chart_data_start+4)
    cats_ref = Reference(ws, min_col=10,
                         min_row=chart_data_start, max_row=chart_data_start+4)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"A{ROW}")


# ── SHEET 4: SENTIMENT ────────────────────────────────────────────────────────

def write_sentiment(wb, df):
    ws = wb.create_sheet("😊 Sentiment")
    ws.sheet_view.showGridLines = False
    banner(ws, "SENTIMENT ANALYSIS", merge_to="G1", sz=14)
    sub_banner(ws, "TextBlob NLP polarity: Positive > 0.1 | Negative < -0.1 | Neutral in between",
               "G", row=2)

    set_col_widths(ws, [24, 10, 10, 12, 12, 14, 14])
    stores = list(df["store"].unique()) + ["ALL STORES"]

    ROW = 4
    for sname in stores:
        sdf = df if sname == "ALL STORES" else df[df["store"] == sname]
        bg  = C["navy"] if sname == "ALL STORES" else C["dkblue"]

        ws.merge_cells(f"A{ROW}:G{ROW}")
        hdr(ws.cell(ROW, 1), sname, bg=bg, sz=11)
        ws.row_dimensions[ROW].height = 26
        ROW += 1

        for ci, h in enumerate(["Sentiment","Count","%","Avg ★","Avg Score","Min ★","Max ★"], 1):
            hdr(ws.cell(ROW, ci), h, bg="365F91", sz=10)
        tbl_s = ROW; ROW += 1

        for label in ["Positive", "Negative", "Neutral", "No Text"]:
            sub    = sdf[sdf["sentiment"] == label]
            cnt    = len(sub)
            pct    = 100 * cnt / max(len(sdf), 1)
            avg_r  = sub["rating"].mean()
            avg_s  = sub["sent_score"].mean()
            min_r  = sub["rating"].min()
            max_r  = sub["rating"].max()
            fill   = SENT_CLR[label]
            vals   = [label, cnt, f"{pct:.1f}%",
                      f"{avg_r:.2f}" if pd.notna(avg_r) else "-",
                      f"{avg_s:.3f}" if pd.notna(avg_s) else "-",
                      f"{int(min_r)}" if pd.notna(min_r) else "-",
                      f"{int(max_r)}" if pd.notna(max_r) else "-"]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(ROW, ci)
                dat(cell, v, center=(ci > 1))
                cell.fill = PatternFill("solid", start_color=fill)
            ROW += 1

        apply_border(ws, tbl_s, ROW-1, 1, 7)
        ROW += 2

    # ── Pie chart — overall sentiment (text reviews only) ──
    pc_col = 9
    pc_row = 4
    ws.cell(pc_row, pc_col).value     = "Sentiment"
    ws.cell(pc_row, pc_col+1).value   = "Count"
    dr = pc_row
    for label in ["Positive", "Negative", "Neutral"]:
        dr += 1
        ws.cell(dr, pc_col).value   = label
        ws.cell(dr, pc_col+1).value = int((df["sentiment"] == label).sum())

    pie = PieChart()
    pie.title  = "Overall Sentiment (text reviews)"
    pie.style  = 10; pie.width = 18; pie.height = 14
    pie.add_data(Reference(ws, min_col=pc_col+1, min_row=pc_row, max_row=dr),
                 titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=pc_col, min_row=pc_row+1, max_row=dr))
    ws.add_chart(pie, f"I20")


# ── SHEET 5: CATEGORIES ───────────────────────────────────────────────────────

def write_categories(wb, df):
    ws = wb.create_sheet("📂 Categories")
    ws.sheet_view.showGridLines = False
    banner(ws, "CATEGORY ANALYSIS — WHAT CUSTOMERS TALK ABOUT", merge_to="I1", sz=14)
    sub_banner(ws, "Keyword-based categorisation. A review may belong to multiple categories.",
               "I", row=2)

    stores = list(df["store"].unique())
    col_hdrs = ["Category", "Total", "Avg ★", "Positive %", "Negative %"] + stores
    set_col_widths(ws, [26, 8, 8, 11, 11] + [22]*len(stores))

    ROW = 4
    for ci, h in enumerate(col_hdrs, 1):
        hdr(ws.cell(ROW, ci), h, bg=C["navy"], sz=10)
    ws.row_dimensions[ROW].height = 24
    tbl_s = ROW; ROW += 1

    all_cats = []
    for cats in df["categories"].dropna():
        all_cats.extend(cats.split(", "))
    cat_total = Counter(all_cats)

    for cat in sorted(cat_total, key=lambda x: -cat_total[x]):
        mask = df["categories"].str.contains(re.escape(cat), na=False)
        cdf  = df[mask]
        pos  = 100*(cdf["sentiment"]=="Positive").sum()/max(len(cdf),1)
        neg  = 100*(cdf["sentiment"]=="Negative").sum()/max(len(cdf),1)
        avg  = cdf["rating"].mean()
        vals = [cat, cat_total[cat],
                f"{avg:.2f}" if pd.notna(avg) else "-",
                f"{pos:.1f}%", f"{neg:.1f}%"]
        for sname in stores:
            vals.append(int(df[(df["store"]==sname) & mask].shape[0]))
        for ci, v in enumerate(vals, 1):
            dat(ws.cell(ROW, ci), v, center=(ci>1), indent=(1 if ci==1 else 0))
            alt(ws, ROW, len(vals))
        ROW += 1

    apply_border(ws, tbl_s, ROW-1, 1, len(col_hdrs))

    # ── Horizontal bar chart ──
    ROW += 2
    n = len(cat_total)
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Review Mentions by Category"
    chart.y_axis.title = "Category"
    chart.x_axis.title = "Number of Mentions"
    chart.style = 10; chart.width = 26; chart.height = 18
    data_ref = Reference(ws, min_col=2, max_col=2,
                         min_row=tbl_s+1, max_row=tbl_s+n)
    cats_ref = Reference(ws, min_col=1,
                         min_row=tbl_s+1, max_row=tbl_s+n)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"A{ROW}")


# ── SHEET 6: TRENDS (REBUILT) ─────────────────────────────────────────────────

def write_trends(wb, df):
    ws = wb.create_sheet("📈 Trends")
    ws.sheet_view.showGridLines = False
    banner(ws, "REVIEW TRENDS OVER TIME", merge_to="L1", sz=14)
    sub_banner(ws,
        "Dates are approximated from Google's relative timestamps (e.g. '3 months ago'). "
        "Current date used as anchor: " + datetime.today().strftime("%d %B %Y"),
        "L", row=2)

    df_t    = df[df["month_year"].str.len() == 7].copy()
    stores  = list(df["store"].unique())

    if df_t.empty:
        ws.cell(4, 1).value = "No date data available — all reviews may lack timestamps."
        return

    # ── SECTION A: YEARLY SUMMARY ─────────────────────────────────────────────
    ROW = 4
    ws.merge_cells(f"A{ROW}:L{ROW}")
    hdr(ws.cell(ROW, 1), "YEARLY SUMMARY", bg=C["dkblue"], sz=11)
    ws.row_dimensions[ROW].height = 26
    ROW += 1

    y_hdrs = ["Year"] + [f"{s}\nReviews" for s in stores] + \
             ["Total\nReviews", "Avg ★", "Positive %", "Negative %"]
    ncols_y = len(y_hdrs)
    for ci, h in enumerate(y_hdrs, 1):
        hdr(ws.cell(ROW, ci), h, bg="365F91", sz=10)
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.row_dimensions[ROW].height = 30
    ytbl_s = ROW; ROW += 1

    years = sorted(df_t["year"].dropna().unique())
    for yr in years:
        ydf = df_t[df_t["year"] == yr]
        avg = ydf["rating"].mean()
        pos = 100*(ydf["sentiment"]=="Positive").sum()/max(len(ydf),1)
        neg = 100*(ydf["sentiment"]=="Negative").sum()/max(len(ydf),1)
        vals = [yr]
        for s in stores:
            vals.append(len(ydf[ydf["store"]==s]))
        vals += [len(ydf),
                 f"{avg:.2f}" if pd.notna(avg) else "-",
                 f"{pos:.1f}%", f"{neg:.1f}%"]
        for ci, v in enumerate(vals, 1):
            dat(ws.cell(ROW, ci), v, center=(ci>1))
            alt(ws, ROW, ncols_y)
        ROW += 1

    apply_border(ws, ytbl_s, ROW-1, 1, ncols_y)

    # ── SECTION B: MONTHLY DETAIL TABLE ──────────────────────────────────────
    ROW += 2
    ws.merge_cells(f"A{ROW}:L{ROW}")
    hdr(ws.cell(ROW, 1), "MONTHLY DETAIL", bg=C["dkblue"], sz=11)
    ws.row_dimensions[ROW].height = 26
    ROW += 1

    m_hdrs = ["Month", "Year"] + [f"{s}\nReviews" for s in stores] + \
             ["Total", "Avg ★", "5★", "4★", "3★", "2★", "1★", "Positive %", "Negative %"]
    for ci, h in enumerate(m_hdrs, 1):
        hdr(ws.cell(ROW, ci), h, bg="365F91", sz=10)
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.row_dimensions[ROW].height = 30
    mtbl_s = ROW; ROW += 1

    months = sorted(df_t["month_year"].unique())
    for month in months:
        mdf  = df_t[df_t["month_year"] == month]
        avg  = mdf["rating"].mean()
        pos  = 100*(mdf["sentiment"]=="Positive").sum()/max(len(mdf),1)
        neg  = 100*(mdf["sentiment"]=="Negative").sum()/max(len(mdf),1)
        rc   = mdf["rating"].value_counts()
        # Display month as "Jan 2024" format
        try:
            label = datetime.strptime(month, "%Y-%m").strftime("%b %Y")
            yr    = datetime.strptime(month, "%Y-%m").strftime("%Y")
        except Exception:
            label = month; yr = month[:4]

        vals = [label, yr]
        for s in stores:
            vals.append(len(mdf[mdf["store"]==s]))
        vals += [len(mdf),
                 f"{avg:.2f}" if pd.notna(avg) else "-",
                 rc.get(5,0), rc.get(4,0), rc.get(3,0), rc.get(2,0), rc.get(1,0),
                 f"{pos:.1f}%", f"{neg:.1f}%"]
        for ci, v in enumerate(vals, 1):
            dat(ws.cell(ROW, ci), v, center=(ci>1))
            alt(ws, ROW, len(vals))
        ROW += 1

    apply_border(ws, mtbl_s, ROW-1, 1, len(m_hdrs))

    # ── CHART 1: Monthly review volume (line chart) ───────────────────────────
    ROW += 2
    n = len(months)
    vol_chart = LineChart()
    vol_chart.title        = "Monthly Review Volume by Store"
    vol_chart.y_axis.title = "Number of Reviews"
    vol_chart.x_axis.title = "Month"
    vol_chart.style        = 10
    vol_chart.width        = 32; vol_chart.height = 16

    for si, s in enumerate(stores):
        col_idx = 3 + si   # cols: Month(1), Year(2), Store1(3), Store2(4), Store3(5)
        data_ref = Reference(ws, min_col=col_idx, max_col=col_idx,
                             min_row=mtbl_s, max_row=mtbl_s + n)
        vol_chart.add_data(data_ref, titles_from_data=True)

    cats_ref = Reference(ws, min_col=1, min_row=mtbl_s+1, max_row=mtbl_s+n)
    vol_chart.set_categories(cats_ref)
    chart_anchor_row = ROW
    ws.add_chart(vol_chart, f"A{chart_anchor_row}")

    # ── CHART 2: Monthly avg rating (line chart) ──────────────────────────────
    # Avg rating is in column = 3 + len(stores) + 1 (after Total col)
    avg_col = 3 + len(stores) + 1

    rat_chart = LineChart()
    rat_chart.title        = "Monthly Average Star Rating (All Stores)"
    rat_chart.y_axis.title = "Avg ★ Rating"
    rat_chart.y_axis.scaling.min = 1
    rat_chart.y_axis.scaling.max = 5
    rat_chart.x_axis.title = "Month"
    rat_chart.style        = 10
    rat_chart.width        = 32; rat_chart.height = 16

    # Build numeric-only avg column for chart (strip the text "f{avg:.2f}")
    chart_avg_col = 20
    ws.cell(mtbl_s, chart_avg_col).value = "Avg Rating (numeric)"
    for i, month in enumerate(months):
        mdf = df_t[df_t["month_year"] == month]
        avg = mdf["rating"].mean()
        ws.cell(mtbl_s + 1 + i, chart_avg_col).value = round(avg, 2) if pd.notna(avg) else None

    data_ref2 = Reference(ws, min_col=chart_avg_col, max_col=chart_avg_col,
                          min_row=mtbl_s, max_row=mtbl_s + n)
    cats_ref2 = Reference(ws, min_col=1, min_row=mtbl_s+1, max_row=mtbl_s+n)
    rat_chart.add_data(data_ref2, titles_from_data=True)
    rat_chart.set_categories(cats_ref2)
    ws.add_chart(rat_chart, f"A{chart_anchor_row + 28}")


# ── SHEET 7: NOTABLE REVIEWS ──────────────────────────────────────────────────

def write_notable(wb, df):
    ws = wb.create_sheet("💬 Notable Reviews")
    ws.sheet_view.showGridLines = False
    banner(ws, "NOTABLE REVIEWS — BEST, WORST & MOST LIKED", merge_to="F1", sz=14)

    set_col_widths(ws, [24, 8, 12, 65, 14, 16])

    def section(title, sub_df, start_row, bg, max_rows=15):
        ws.merge_cells(f"A{start_row}:F{start_row}")
        hdr(ws.cell(start_row, 1), title, bg=bg, sz=11)
        ws.row_dimensions[start_row].height = 26
        start_row += 1
        for ci, h in enumerate(["Store","★","Sentiment","Review Text","Date","Author"], 1):
            hdr(ws.cell(start_row, ci), h, bg="365F91", sz=10)
        start_row += 1
        shown = sub_df.head(max_rows)
        for _, r in shown.iterrows():
            ws.cell(start_row, 1).value = r["store"]
            ws.cell(start_row, 2).value = r["rating"]
            ws.cell(start_row, 3).value = r["sentiment"]
            ws.cell(start_row, 4).value = str(r["text"])[:500]
            ws.cell(start_row, 5).value = r["date_raw"]
            ws.cell(start_row, 6).value = r["author"]
            for ci in range(1, 7):
                cell = ws.cell(start_row, ci)
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=(ci==4))
            snt = str(r["sentiment"])
            if snt in SENT_CLR:
                ws.cell(start_row, 3).fill = PatternFill("solid", start_color=SENT_CLR[snt])
            rat = r["rating"]
            if pd.notna(rat) and int(rat) in STAR_CLR:
                ws.cell(start_row, 2).fill = PatternFill("solid", start_color=STAR_CLR[int(rat)])
                ws.cell(start_row, 2).alignment = Alignment(horizontal="center", vertical="top")
            start_row += 1
        apply_border(ws, start_row - len(shown) - 1, start_row-1, 1, 6)
        return start_row + 2

    ROW = 3
    text_df  = df[df["text"].str.strip().astype(bool)]
    top5     = text_df[text_df["rating"] == 5].sample(frac=1, random_state=1)
    bottom12 = text_df[text_df["rating"].isin([1, 2])].sort_values("likes", ascending=False)
    most_lkd = df.sort_values("likes", ascending=False)

    ROW = section("⭐ BEST REVIEWS (5 Stars, sample)",        top5,    ROW, C["green"])
    ROW = section("⚠  CRITICAL REVIEWS (1–2 Stars)",         bottom12, ROW, C["red"])
    ROW = section("👍 MOST HELPFUL / LIKED REVIEWS",          most_lkd, ROW, C["dkblue"])

    # ── Per-store worst reviews ──
    stores = list(df["store"].unique())
    for sname in stores:
        worst = text_df[(text_df["store"]==sname) &
                        (text_df["rating"].isin([1,2]))].sort_values("likes", ascending=False)
        ROW = section(f"⚠  CRITICAL — {sname}", worst, ROW, "7B2C2C", max_rows=10)


# ── SHEET 8: WORD FREQUENCY ───────────────────────────────────────────────────

STOPWORDS = set("""
a an the and or but in on at to for of is are was were be been being
have has had do does did will would could should may might shall can
i me my we our you your he she it its they them their this that these
those with from by not no so if as when then there also very just
really good great nice well s t""".split())

def top_words(texts, n=30):
    words = []
    for t in texts:
        if not isinstance(t, str): continue
        for w in re.findall(r"[a-z]{3,}", t.lower()):
            if w not in STOPWORDS:
                words.append(w)
    return Counter(words).most_common(n)

def write_words(wb, df):
    ws = wb.create_sheet("🔤 Top Words")
    ws.sheet_view.showGridLines = False
    banner(ws, "TOP WORDS IN REVIEWS (by Sentiment & Store)", merge_to="K1", sz=14)
    sub_banner(ws, "Common stopwords removed. Minimum word length: 3 characters.", "K", row=2)

    text_df = df[df["text"].str.strip().astype(bool)]
    sections_data = [
        ("ALL — Positive Reviews",  text_df[text_df["sentiment"]=="Positive"]["text"], C["green"]),
        ("ALL — Negative Reviews",  text_df[text_df["sentiment"]=="Negative"]["text"], C["red"]),
        ("ALL — Neutral Reviews",   text_df[text_df["sentiment"]=="Neutral"]["text"],  "7D6608"),
    ]
    for sname in df["store"].unique():
        sdf = text_df[text_df["store"]==sname]
        sections_data.append((f"{sname} — All Text", sdf["text"], C["dkblue"]))

    ROW = 4
    col = 1
    for title, texts, bg in sections_data:
        words = top_words(texts, n=25)
        if not words: continue

        ws.merge_cells(start_row=ROW, start_column=col, end_row=ROW, end_column=col+1)
        hdr(ws.cell(ROW, col), title, bg=bg, sz=10)
        ws.row_dimensions[ROW].height = 24

        hdr(ws.cell(ROW+1, col),   "Word",  bg="365F91", sz=9)
        hdr(ws.cell(ROW+1, col+1), "Count", bg="365F91", sz=9)
        ws.column_dimensions[get_column_letter(col)].width   = 18
        ws.column_dimensions[get_column_letter(col+1)].width = 8

        for i, (word, cnt) in enumerate(words):
            ws.cell(ROW+2+i, col).value   = word
            ws.cell(ROW+2+i, col+1).value = cnt
            ws.cell(ROW+2+i, col).font    = Font(name="Arial", size=9)
            ws.cell(ROW+2+i, col+1).font  = Font(name="Arial", size=9)
            ws.cell(ROW+2+i, col+1).alignment = Alignment(horizontal="center")
            alt(ws, ROW+2+i, 2, c1=col)

        apply_border(ws, ROW, ROW+1+len(words), col, col+1)
        col += 3   # move to next section
        if col > 10:   # wrap to next block of rows
            col = 1
            ROW += 30


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    if not Path(INPUT_JSON).exists():
        print(f"ERROR: {INPUT_JSON} not found. Run the scraper first.")
        return

    with open(INPUT_JSON, "r", encoding="utf-8") as f:
        raw = json.load(f)

    print(f"Loaded {len(raw)} reviews from {INPUT_JSON}")
    df = build_df(raw)

    dated = df[df["month_year"].str.len()==7]
    print(f"Reviews with dates parsed: {len(dated)} / {len(df)}")
    print(f"Years covered: {sorted(df['year'].dropna().unique())}")
    print(f"Stores: {list(df['store'].unique())}")

    wb = Workbook()
    wb.remove(wb.active)

    steps = [
        ("Dashboard",           write_dashboard),
        ("All Reviews",         write_raw),
        ("Ratings",             write_ratings),
        ("Sentiment",           write_sentiment),
        ("Categories",          write_categories),
        ("Trends",              write_trends),
        ("Notable Reviews",     write_notable),
        ("Top Words",           write_words),
    ]
    for name, fn in steps:
        print(f"  -> Writing {name}...")
        fn(wb, df)

    wb.save(OUTPUT_XLSX)
    print(f"\n{'='*60}")
    print(f"  ✓ Saved: {OUTPUT_XLSX}")
    print(f"  ✓ Sheets: " + " | ".join(n for n, _ in steps))
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()