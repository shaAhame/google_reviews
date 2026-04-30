"""
╔══════════════════════════════════════════════════════════════════════╗
║   IDEALZ — WEEKLY REVIEW TRACKER                                     ║
║   Run every Monday morning                                           ║
║   Generates PDF + Excel report — you send it manually to CEO         ║
╚══════════════════════════════════════════════════════════════════════╝

HOW IT WORKS:
  - Scrapes latest 200 reviews per store (sorted Newest first)
  - Saves snapshot: snapshots/reviews_YYYY-MM-DD.json
  - Compares with last Monday's snapshot
  - New IDs this week = reviews posted this week
  - Generates: weekly_report_YYYY-MM-DD.pdf
               weekly_report_YYYY-MM-DD.xlsx

IMPORTANT:
  - Run EVERY Monday without skipping
  - First run = baseline only (no report yet)
  - Second run next Monday = first real report

SETUP (one time):
    pip install playwright textblob openpyxl pandas reportlab
    playwright install chromium

RUN:
    python idealz_weekly_tracker.py
"""

import json
import re
import sys
from pathlib import Path
from datetime import datetime
from collections import Counter

import pandas as pd
from textblob import TextBlob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, PageBreak, HRFlowable)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# ══════════════════════════════════════════════════════════════════════
# CONFIG — Only thing you need to edit
# ══════════════════════════════════════════════════════════════════════

STORES = [
    {
        "name": "Idealz Prime",
        "url":  "https://www.google.com/maps/place/iDealz+Prime/@6.8912695,79.8560961,17z/data=!3m1!4b1!4m6!3m5!1s0x3ae259005a2260c1:0xd6febd8ffeac3a34!8m2!3d6.8912695!4d79.8560961!16s%2Fg%2F11w27bncwk?entry=ttu",
    },
    {
        "name": "Idealz Lanka - Marino Mall",
        "url":  "https://www.google.com/maps/place/iDealz+Lanka+-+Marino+Mall/@6.9001796,79.8523305,17z/data=!3m1!4b1!4m6!3m5!1s0x3ae25957ebf8012b:0xe0e160f3a83edd3c!8m2!3d6.9001796!4d79.8523305!16s%2Fg%2F11gr41k7q8?entry=ttu",
    },
    {
        "name": "Idealz Lanka - Liberty Plaza",
        "url":  "https://www.google.com/maps/place/iDealz+Lanka+Pvt+Ltd/@6.9116839,79.8515051,17z/data=!3m1!4b1!4m6!3m5!1s0x3ae25911b0316acb:0xdd0d30f303baddf1!8m2!3d6.9116839!4d79.8515051!16s%2Fg%2F11b6dg62r6?entry=ttu",
    },
]

# Reviews to scrape per store — 200 safely covers any week
SCRAPE_LIMIT = 200

# ══════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════

SNAPSHOT_DIR = Path("snapshots")
TODAY        = datetime.today().strftime("%Y-%m-%d")
OUTPUT_XLSX  = f"weekly_report_{TODAY}.xlsx"
OUTPUT_PDF   = f"weekly_report_{TODAY}.pdf"

CATEGORIES = {
    "Product Quality":  ["quality","product","item","defect","broken","material",
                         "durable","genuine","fake","original","condition","damaged"],
    "Customer Service": ["service","staff","employee","helpful","rude","assist",
                         "support","polite","attitude","friendly","ignored",
                         "customer care","respond","behavior","behaviour"],
    "Pricing & Value":  ["price","pricing","expensive","cheap","affordable","value",
                         "worth","overpriced","discount","offer","deal","cost","money"],
    "Store Experience": ["store","shop","location","parking","clean","organized",
                         "display","ambiance","atmosphere","interior","mall","crowded",
                         "visit","branch","outlet"],
    "After-Sales":      ["return","refund","exchange","warranty","replace",
                         "complaint","issue","resolve","compensation","claim"],
    "Variety / Stock":  ["variety","selection","stock","available","range",
                         "choice","collection","assortment","limited","out of stock"],
}

# ══════════════════════════════════════════════════════════════════════
# SCRAPER
# ══════════════════════════════════════════════════════════════════════

def scrape_all_stores():
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  Run: pip install playwright && playwright install chromium")
        sys.exit(1)

    all_reviews = []
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False, slow_mo=30,
            args=["--disable-blink-features=AutomationControlled",
                  "--no-sandbox", "--start-maximized"]
        )
        ctx = browser.new_context(
            viewport={"width": 1400, "height": 900},
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36"),
            locale="en-US",
            timezone_id="Asia/Colombo",
        )
        page = ctx.new_page()

        # Accept cookie consent
        page.goto("https://www.google.com/maps", wait_until="domcontentloaded")
        page.wait_for_timeout(2000)
        for sel in ['button[aria-label*="Accept"]', 'button:has-text("Accept all")']:
            try:
                btn = page.query_selector(sel)
                if btn:
                    btn.click()
                    page.wait_for_timeout(800)
                    break
            except:
                pass

        for store in STORES:
            print(f"\n  Scraping: {store['name']}")
            try:
                reviews = _scrape_store(page, store)
                print(f"  [OK] {len(reviews)} reviews collected")
                all_reviews.extend(reviews)
            except Exception as e:
                print(f"  [X] Failed: {e}")
            page.wait_for_timeout(3000)

        browser.close()
    return all_reviews


def _scrape_store(page, store):
    url = store["url"] + ("&" if "?" in store["url"] else "?") + "hl=en"
    page.goto(url, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(3000)

    # Go to place page if redirected to search
    if "/maps/place/" not in page.url:
        for sel in ['a[href*="/maps/place/"]', '.hfpxzc']:
            el = page.query_selector(sel)
            if el:
                el.click()
                page.wait_for_load_state("domcontentloaded")
                page.wait_for_timeout(4000)
                break

    if "/maps/place/" not in page.url:
        raise Exception("Could not navigate to place page")

    # Click Reviews tab
    for sel in ['button[aria-label*="Reviews"]', 'button[role="tab"]:has-text("Reviews")']:
        el = page.query_selector(sel)
        if el:
            el.click()
            page.wait_for_timeout(3000)
            break

    # Sort by Newest
    for sel in ['button[aria-label^="Sort"]', 'button[data-value="Sort"]']:
        btn = page.query_selector(sel)
        if btn:
            btn.click()
            page.wait_for_timeout(1500)
            for opt in ['div[data-index="1"]', 'li:has-text("Newest")']:
                o = page.query_selector(opt)
                if o:
                    o.click()
                    page.wait_for_timeout(2500)
                    break
            break

    try:
        page.wait_for_selector('[data-review-id]', timeout=12000)
    except:
        return []

    page.wait_for_timeout(2000)

    # Scroll until we have SCRAPE_LIMIT reviews
    seen_ids = set()
    stall    = 0
    while stall < 8:
        page.evaluate("""() => {
            const sels = ['div[role="feed"]','.m6QErb.DxyBCb','.m6QErb[aria-label]','.m6QErb'];
            for (const s of sels) {
                const el = document.querySelector(s);
                if (el && el.scrollHeight > el.clientHeight) { el.scrollTop += 5000; return; }
            }
        }""")
        page.wait_for_timeout(2500)

        ids      = set(page.evaluate("""() =>
            Array.from(document.querySelectorAll('[data-review-id]'))
                 .map(el => el.getAttribute('data-review-id')).filter(Boolean)
        """))
        new_ids  = ids - seen_ids
        seen_ids = ids
        print(f"    {len(seen_ids)} / {SCRAPE_LIMIT} loaded...", end="\r", flush=True)

        if len(seen_ids) >= SCRAPE_LIMIT:
            print(f"\n    [OK] Reached {SCRAPE_LIMIT}")
            break
        stall = 0 if new_ids else stall + 1

    print(f"\n    {len(seen_ids)} unique reviews in DOM")

    # Expand truncated review text
    for btn in page.query_selector_all('button.w8nwRe, button[aria-label="See more"]'):
        try:
            btn.click()
            page.wait_for_timeout(50)
        except:
            pass

    # Extract review data
    cards     = page.query_selector_all('[data-review-id]')
    reviews   = []
    extracted = set()

    for card in cards:
        try:
            rid = card.get_attribute("data-review-id") or ""
            if rid in extracted:
                continue
            extracted.add(rid)

            author = "Anonymous"
            for s in ['.d4r55', '.DUwDvf']:
                el = card.query_selector(s)
                if el and el.inner_text().strip():
                    author = el.inner_text().strip()
                    break

            rating = None
            for s in ['[aria-label*="star"]', '.kvMYJc']:
                el = card.query_selector(s)
                if el:
                    m = re.search(r'(\d+)', el.get_attribute("aria-label") or "")
                    if m:
                        rating = int(m.group(1))
                        break

            text = ""
            for s in ['.wiI7pd', '.MyEned']:
                el = card.query_selector(s)
                if el and el.inner_text().strip():
                    text = el.inner_text().strip()
                    break

            date_str = ""
            for s in ['.rsqaWe', '.dehysf', '.xRkPPb']:
                el = card.query_selector(s)
                if el and el.inner_text().strip():
                    date_str = el.inner_text().strip()
                    break

            owner_reply = ""
            el = card.query_selector('.CDe7pd')
            if el:
                owner_reply = el.inner_text().strip()

            reviews.append({
                "review_id":   rid,
                "store":       store["name"],
                "author":      author,
                "rating":      rating,
                "date_raw":    date_str,
                "text":        text,
                "owner_reply": owner_reply,
                "scraped_on":  TODAY,
                "sentiment":   "",
                "sent_score":  0.0,
                "categories":  "",
            })
        except:
            continue

    return reviews[:SCRAPE_LIMIT]


# ══════════════════════════════════════════════════════════════════════
# SNAPSHOT MANAGEMENT
# ══════════════════════════════════════════════════════════════════════

def save_snapshot(reviews):
    SNAPSHOT_DIR.mkdir(exist_ok=True)
    path = SNAPSHOT_DIR / f"reviews_{TODAY}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(reviews, f, ensure_ascii=False, indent=2)
    print(f"  [OK] Snapshot saved -> {path}")


def load_previous_snapshot():
    """Load the most recent snapshot before today."""
    if not SNAPSHOT_DIR.exists():
        return None, None
    files    = sorted(SNAPSHOT_DIR.glob("reviews_*.json"))
    prev     = [f for f in files if f.stem != f"reviews_{TODAY}"]
    if not prev:
        return None, None
    latest   = prev[-1]
    date_str = latest.stem.replace("reviews_", "")
    with open(latest, "r", encoding="utf-8") as f:
        data = json.load(f)
    print(f"  [OK] Previous snapshot: {latest.name}  ({len(data)} reviews)")
    return data, date_str


def load_all_snapshots():
    if not SNAPSHOT_DIR.exists():
        return {}
    result = {}
    for f in sorted(SNAPSHOT_DIR.glob("reviews_*.json")):
        date_str = f.stem.replace("reviews_", "")
        with open(f, "r", encoding="utf-8") as fh:
            result[date_str] = json.load(fh)
    return result


# ══════════════════════════════════════════════════════════════════════
# ANALYSIS
# ══════════════════════════════════════════════════════════════════════

def get_sentiment(text):
    if not text or not str(text).strip():
        return "No Text", 0.0
    try:
        score = TextBlob(str(text)).sentiment.polarity
    except:
        score = 0.0
    if score > 0.1:  return "Positive", round(score, 3)
    if score < -0.1: return "Negative", round(score, 3)
    return "Neutral", round(score, 3)


def categorize(text):
    if not text or not isinstance(text, str):
        return "General"
    tl      = text.lower()
    matched = [cat for cat, kws in CATEGORIES.items() if any(kw in tl for kw in kws)]
    return ", ".join(matched) if matched else "General"


def enrich(reviews):
    """Add sentiment + category. Skips already-enriched rows."""
    for r in reviews:
        if not r.get("sentiment"):
            r["sentiment"], r["sent_score"] = get_sentiment(r.get("text", ""))
        if not r.get("categories"):
            r["categories"] = categorize(r.get("text", ""))
    return reviews


def get_new_reviews(current, previous):
    """
    Core weekly diff.
    Returns reviews in current whose review_id was NOT in previous snapshot.
    These are the reviews posted THIS WEEK.
    """
    if not previous:
        return []
    prev_ids = {r.get("review_id", "") for r in previous}
    return [r for r in current if r.get("review_id", "") not in prev_ids]


def weekly_stats(new_reviews):
    """Compute all report metrics from this week's new reviews only."""
    EMPTY = {
        "total": 0, "critical": 0, "low": 0, "good": 0,
        "positive": 0, "negative": 0, "neutral": 0, "no_text": 0,
        "with_reply": 0, "without_reply": 0, "avg_rating": None,
        "star_dist": {5:0, 4:0, 3:0, 2:0, 1:0},
        "by_store":  {s["name"]: {"total":0,"critical":0,"avg":None,
                                   "positive":0,"negative":0} for s in STORES},
        "categories": Counter(), "critical_list": [], "unanswered_neg": [],
    }
    if not new_reviews:
        return EMPTY

    df           = pd.DataFrame(new_reviews)
    df["rating"] = pd.to_numeric(df.get("rating"), errors="coerce")
    total        = len(df)
    avg          = df["rating"].mean()

    critical = int((df["rating"] <= 2).sum())
    low      = int((df["rating"] == 3).sum())
    good     = int((df["rating"] >= 4).sum())

    sent     = df.get("sentiment", pd.Series(["No Text"] * total))
    positive = int((sent == "Positive").sum())
    negative = int((sent == "Negative").sum())
    neutral  = int((sent == "Neutral").sum())
    no_text  = int((sent == "No Text").sum())

    reply_col     = df.get("owner_reply", pd.Series([""] * total))
    with_reply    = int(reply_col.astype(str).str.strip().astype(bool).sum())
    without_reply = total - with_reply

    star_dist = {s: int((df["rating"] == s).sum()) for s in [5, 4, 3, 2, 1]}

    by_store = {}
    for s in STORES:
        sname = s["name"]
        sr    = [r for r in new_reviews if r["store"] == sname]
        sdf   = pd.DataFrame(sr) if sr else pd.DataFrame()
        by_store[sname] = {
            "total":    len(sr),
            "critical": int((pd.to_numeric(sdf.get("rating", pd.Series()), errors="coerce") <= 2).sum()) if sr else 0,
            "avg":      round(pd.to_numeric(sdf.get("rating", pd.Series()), errors="coerce").mean(), 2) if sr else None,
            "positive": int((sdf.get("sentiment", pd.Series()) == "Positive").sum()) if sr else 0,
            "negative": int((sdf.get("sentiment", pd.Series()) == "Negative").sum()) if sr else 0,
        }

    cat_counter = Counter()
    for r in new_reviews:
        for cat in r.get("categories", "General").split(", "):
            cat_counter[cat.strip()] += 1

    critical_list  = sorted(
        [r for r in new_reviews if (r.get("rating") or 5) <= 2],
        key=lambda r: r.get("rating") or 0
    )
    unanswered_neg = [
        r for r in new_reviews
        if ((r.get("rating") or 5) <= 3 or r.get("sentiment") == "Negative")
        and not r.get("owner_reply", "").strip()
    ]

    return {
        "total": total, "critical": critical, "low": low, "good": good,
        "positive": positive, "negative": negative, "neutral": neutral, "no_text": no_text,
        "with_reply": with_reply, "without_reply": without_reply,
        "avg_rating": round(avg, 2) if pd.notna(avg) else None,
        "star_dist": star_dist, "by_store": by_store, "categories": cat_counter,
        "critical_list": critical_list, "unanswered_neg": unanswered_neg,
    }


# ══════════════════════════════════════════════════════════════════════
# EXCEL HELPERS
# ══════════════════════════════════════════════════════════════════════

def _side():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _borders(ws, r1, r2, c1, c2):
    b = _side()
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cell in row:
            cell.border = b

def _hdr(cell, text, bg="1F3864", fg="FFFFFF", sz=10, bold=True):
    cell.value     = text
    cell.font      = Font(name="Arial", bold=bold, size=sz, color=fg)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _side()

def _dat(cell, value, center=False, sz=10, fg="000000", bold=False, bg=None):
    cell.value     = value
    cell.font      = Font(name="Arial", size=sz, color=fg, bold=bold)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=True)
    cell.border    = _side()
    if bg:
        cell.fill  = PatternFill("solid", start_color=bg)

def _banner(ws, text, end_col, row=1, bg="1F3864", sz=13):
    ws.merge_cells(f"A{row}:{end_col}{row}")
    _hdr(ws[f"A{row}"], text, bg=bg, sz=sz)
    ws.row_dimensions[row].height = 36


# ══════════════════════════════════════════════════════════════════════
# EXCEL REPORT
# ══════════════════════════════════════════════════════════════════════

def build_excel_report(stats, new_reviews, prev_date, all_snapshots):
    wb = Workbook()
    wb.remove(wb.active)
    _xl_summary(wb, stats, prev_date)
    _xl_all_new(wb, new_reviews)
    _xl_critical(wb, stats)
    _xl_trends(wb, all_snapshots)
    wb.save(OUTPUT_XLSX)
    print(f"  [OK] Excel saved -> {OUTPUT_XLSX}")


def _xl_summary(wb, stats, prev_date):
    ws = wb.create_sheet("Weekly Summary")
    ws.sheet_view.showGridLines = False
    _banner(ws, f"IDEALZ — WEEKLY REVIEW SUMMARY  |  Week ending: {TODAY}", "L")

    ws.merge_cells("A2:L2")
    ws["A2"].value     = (f"vs. snapshot from: {prev_date}   |   "
                          f"New reviews this week: {stats['total']}   |   "
                          f"Critical (1-2 stars): {stats['critical']}   |   "
                          f"Without reply: {stats['without_reply']}")
    ws["A2"].font      = Font(name="Arial", size=10, italic=True, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # ── Section A: Overall counts ──
    ROW = 4
    ws.merge_cells(f"A{ROW}:N{ROW}")
    _hdr(ws.cell(ROW, 1), "THIS WEEK AT A GLANCE", bg="2E75B6", sz=11)
    ws.row_dimensions[ROW].height = 22
    ROW += 1

    kpi_labels = ["Total\nNew Reviews", "Critical\n(1-2 stars)",
                  "Needs Att.\n(3 stars)", "Positive\n(4-5 stars)",
                  "Positive\nSentiment",  "Negative\nSentiment",
                  "No Owner\nReply"]
    kpi_values = [stats["total"],   stats["critical"],  stats["low"],
                  stats["good"],    stats["positive"],  stats["negative"],
                  stats["without_reply"]]
    kpi_bgs    = ["DDEBF7", "FFC7CE", "FFEB9C",
                  "C6EFCE", "C6EFCE", "FFC7CE", "FCE4D6"]

    for ki in range(len(kpi_labels)):
        c  = ki + 1
        r1 = ROW;  r2 = ROW + 1
        ws.merge_cells(start_row=r1, start_column=c*2-1, end_row=r1, end_column=c*2)
        ws.merge_cells(start_row=r2, start_column=c*2-1, end_row=r2, end_column=c*2)
        _hdr(ws.cell(r1, c*2-1), kpi_labels[ki], bg="365F91", sz=9)
        _hdr(ws.cell(r2, c*2-1), str(kpi_values[ki]),
             bg=kpi_bgs[ki], fg="1A1A1A", sz=16, bold=True)
        ws.row_dimensions[r1].height = 26
        ws.row_dimensions[r2].height = 34
    ROW += 3

    # ── Section B: Per-store table ──
    ws.merge_cells(f"A{ROW}:H{ROW}")
    _hdr(ws.cell(ROW, 1), "PER STORE BREAKDOWN  (This Week Only)", bg="2E75B6", sz=11)
    ws.row_dimensions[ROW].height = 22
    ROW += 1

    hdrs   = ["Branch", "New\nReviews", "Critical\n(1-2 stars)",
              "Needs Att.\n(3 stars)", "Positive\n(4-5 stars)",
              "Avg Rating", "Positive\nSentiment", "Negative\nSentiment"]
    widths = [28, 10, 14, 14, 13, 11, 14, 14]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws.cell(ROW, ci), h, bg="1A4F8A", sz=9)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[ROW].height = 28
    tbl_s = ROW;  ROW += 1

    for si, s in enumerate(STORES):
        sname = s["name"]
        sd    = stats["by_store"].get(sname, {})
        tot   = sd.get("total", 0)
        crit  = sd.get("critical", 0)
        low_s = len([r for r in stats["critical_list"]
                     if r["store"] == sname and (r.get("rating") or 0) == 3])
        good  = tot - crit - low_s
        avg   = sd.get("avg")
        pos_s = sd.get("positive", 0)
        neg_s = sd.get("negative", 0)
        bg    = "F0F4F8" if si % 2 == 0 else "FFFFFF"
        vals  = [sname, tot, crit, low_s, good, avg or "N/A", pos_s, neg_s]
        for ci, v in enumerate(vals, 1):
            cell_bg = bg
            if ci == 3 and crit > 0:  cell_bg = "FFC7CE"
            if ci == 4 and low_s > 0: cell_bg = "FFEB9C"
            if ci == 5 and good > 0:  cell_bg = "C6EFCE"
            _dat(ws.cell(ROW, ci), v, center=(ci > 1), bg=cell_bg)
        ROW += 1
    _borders(ws, tbl_s, ROW-1, 1, len(hdrs))
    ROW += 1

    # ── Section C: Star distribution ──
    ws.merge_cells(f"A{ROW}:H{ROW}")
    _hdr(ws.cell(ROW, 1), "STAR DISTRIBUTION  (New Reviews This Week)", bg="2E75B6", sz=11)
    ws.row_dimensions[ROW].height = 22
    ROW += 1

    star_bg = {5:"C6EFCE", 4:"DDEBF7", 3:"FFF2CC", 2:"FCE4D6", 1:"FFC7CE"}
    total   = stats["total"] or 1
    for star in [5, 4, 3, 2, 1]:
        cnt = stats["star_dist"].get(star, 0)
        pct = round(100 * cnt / total, 1)
        bar = chr(9608) * int(pct / 5)
        ws.merge_cells(f"A{ROW}:B{ROW}")
        _hdr(ws.cell(ROW, 1), f"{star} stars", bg=star_bg[star], fg="000000", sz=11)
        ws.merge_cells(f"C{ROW}:D{ROW}")
        _dat(ws.cell(ROW, 3), f"{cnt}  ({pct}%)", center=True, bold=True)
        ws.merge_cells(f"E{ROW}:H{ROW}")
        ws.cell(ROW, 5).value  = bar + f"  {pct}%"
        ws.cell(ROW, 5).font   = Font(name="Arial", size=10,
            color="1E5631" if star >= 4 else ("9C0006" if star <= 2 else "7D6608"))
        ws.cell(ROW, 5).fill   = PatternFill("solid", start_color=star_bg[star])
        ws.cell(ROW, 5).border = _side()
        ws.row_dimensions[ROW].height = 22
        ROW += 1
    ROW += 1

    # ── Section D: Categories ──
    ws.merge_cells(f"A{ROW}:H{ROW}")
    _hdr(ws.cell(ROW, 1), "TOPICS MENTIONED  (New Reviews This Week)", bg="2E75B6", sz=11)
    ws.row_dimensions[ROW].height = 22
    ROW += 1
    for ci, (h, w) in enumerate(zip(["Category", "Count", "% of Reviews"], [26, 10, 14]), 1):
        _hdr(ws.cell(ROW, ci), h, bg="1A4F8A", sz=9)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[ROW].height = 22
    tbl_s2 = ROW;  ROW += 1
    for cat, cnt in stats["categories"].most_common():
        pct = round(100 * cnt / (stats["total"] or 1), 1)
        _dat(ws.cell(ROW, 1), cat)
        _dat(ws.cell(ROW, 2), cnt,        center=True)
        _dat(ws.cell(ROW, 3), f"{pct}%",  center=True)
        ROW += 1
    _borders(ws, tbl_s2, ROW-1, 1, 3)


def _xl_all_new(wb, new_reviews):
    ws = wb.create_sheet("All New Reviews")
    ws.sheet_view.showGridLines = False
    _banner(ws, f"ALL NEW REVIEWS THIS WEEK  ({len(new_reviews)} total)", "H")

    if not new_reviews:
        ws.merge_cells("A3:H3")
        ws["A3"].value     = "No new reviews this week."
        ws["A3"].font      = Font(name="Arial", size=11, italic=True, color="888888")
        ws["A3"].alignment = Alignment(horizontal="center")
        return

    hdrs   = ["Store", "Stars", "Sentiment", "Review Text", "Date", "Author", "Category", "Replied?"]
    widths = [24, 5, 11, 65, 14, 18, 28, 9]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws.cell(2, ci), h, sz=9)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    SENT_BG = {"Positive":"C6EFCE","Negative":"FFC7CE","Neutral":"FFEB9C","No Text":"F2F2F2"}
    STAR_BG = {5:"C6EFCE", 4:"DDEBF7", 3:"FFF2CC", 2:"FCE4D6", 1:"FFC7CE"}

    sorted_rev = sorted(
        new_reviews,
        key=lambda r: (r.get("sentiment","") != "Negative", -(r.get("rating") or 3))
    )
    for ri, r in enumerate(sorted_rev, 3):
        replied = "Yes" if r.get("owner_reply","").strip() else "No"
        vals    = [r.get("store",""), r.get("rating",""), r.get("sentiment",""),
                   r.get("text","")[:500], r.get("date_raw",""), r.get("author",""),
                   r.get("categories",""), replied]
        for ci, v in enumerate(vals, 1):
            cell           = ws.cell(ri, ci)
            cell.value     = v
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == 4))
            cell.border    = _side()
        sent = r.get("sentiment","")
        if sent in SENT_BG:
            ws.cell(ri, 3).fill = PatternFill("solid", start_color=SENT_BG[sent])
        rat = r.get("rating")
        if rat and int(rat) in STAR_BG:
            ws.cell(ri, 2).fill      = PatternFill("solid", start_color=STAR_BG[int(rat)])
            ws.cell(ri, 2).alignment = Alignment(horizontal="center", vertical="top")

    ws.freeze_panes    = "A3"
    ws.auto_filter.ref = "A2:H2"


def _xl_critical(wb, stats):
    ws = wb.create_sheet("Critical Reviews")
    ws.sheet_view.showGridLines = False
    _banner(ws,
            f"CRITICAL & UNANSWERED  |  {stats['critical']} critical  |  "
            f"{len(stats['unanswered_neg'])} unanswered",
            "G", bg="9C0006")

    if not stats["critical_list"] and not stats["unanswered_neg"]:
        ws.merge_cells("A3:G3")
        ws.cell(3,1).value     = "No critical reviews this week! Great job."
        ws.cell(3,1).font      = Font(name="Arial", size=13, bold=True, color="1E5631")
        ws.cell(3,1).alignment = Alignment(horizontal="center")
        ws.cell(3,1).fill      = PatternFill("solid", start_color="C6EFCE")
        return

    # Combine critical + unanswered, deduplicated
    all_flagged = list({r["review_id"]: r
                        for r in (stats["critical_list"] + stats["unanswered_neg"])}.values())
    all_flagged.sort(key=lambda r: r.get("rating") or 5)

    ROW    = 3
    hdrs   = ["Store", "Stars", "Review Text", "Date", "Author", "Category", "Replied?"]
    widths = [24, 5, 70, 14, 18, 28, 9]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws.cell(ROW, ci), h, bg="1A4F8A", sz=9)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[ROW].height = 22
    tbl_s = ROW;  ROW += 1

    for r in all_flagged:
        rat     = r.get("rating") or 0
        replied = r.get("owner_reply","").strip()
        row_bg  = "FFC7CE" if rat <= 2 else "FFEB9C"
        vals    = [r.get("store",""), rat, r.get("text","")[:600],
                   r.get("date_raw",""), r.get("author",""), r.get("categories",""),
                   "Yes" if replied else "NO REPLY"]
        for ci, v in enumerate(vals, 1):
            cell           = ws.cell(ROW, ci)
            cell.value     = v
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == 3))
            cell.border    = _side()
            cell.fill      = PatternFill("solid", start_color=row_bg)
        # Reply cell: red if no reply, green if replied
        ws.cell(ROW, 7).fill = PatternFill("solid",
            start_color="C6EFCE" if replied else "FF0000")
        ws.cell(ROW, 7).font = Font(name="Arial", size=9,
            color="FFFFFF" if not replied else "000000", bold=(not replied))
        ROW += 1
    _borders(ws, tbl_s, ROW-1, 1, 7)


def _xl_trends(wb, all_snapshots):
    ws = wb.create_sheet("Weekly Trends")
    ws.sheet_view.showGridLines = False
    _banner(ws, "WEEK-ON-WEEK NEW REVIEW TRENDS", "J")

    dates = sorted(all_snapshots.keys())
    if len(dates) < 2:
        ws.cell(3,1).value = "Trend data will appear after 2+ weekly runs."
        return

    rows     = []
    prev_ids = set()
    for date_str in dates:
        snap     = all_snapshots[date_str]
        enrich(snap)
        curr_ids = {r.get("review_id","") for r in snap}
        new      = [r for r in snap if r.get("review_id","") not in prev_ids]
        enrich(new)
        prev_ids = curr_ids
        for s in STORES:
            sname = s["name"]
            s_new = [r for r in new if r["store"] == sname]
            sdf   = pd.DataFrame(s_new) if s_new else pd.DataFrame()
            rows.append({
                "date":     date_str,
                "store":    sname,
                "new":      len(s_new),
                "critical": int((pd.to_numeric(sdf.get("rating", pd.Series()),
                                               errors="coerce") <= 2).sum()) if s_new else 0,
                "avg":      round(pd.to_numeric(sdf.get("rating", pd.Series()),
                                                errors="coerce").mean(), 2) if s_new else None,
            })

    df          = pd.DataFrame(rows)
    store_names = [s["name"] for s in STORES]
    all_dates   = sorted(df["date"].unique())

    ROW = 3
    th  = (["Week"] +
           [f"{s}\nNew" for s in store_names] +
           [f"{s}\nCritical" for s in store_names] +
           [f"{s}\nAvg Rating" for s in store_names])
    for ci, h in enumerate(th, 1):
        _hdr(ws.cell(ROW, ci), h, sz=9)
        ws.column_dimensions[get_column_letter(ci)].width = 15
    ws.row_dimensions[ROW].height = 28
    tbl_s = ROW;  ROW += 1

    for date_str in all_dates:
        ddf  = df[df["date"] == date_str]
        vals = [date_str]
        for col in ["new", "critical", "avg"]:
            for sname in store_names:
                row_s = ddf[ddf["store"] == sname]
                v     = row_s[col].values[0] if len(row_s) else ""
                vals.append(v)
        for ci, v in enumerate(vals, 1):
            _dat(ws.cell(ROW, ci), v, center=(ci > 1))
        ROW += 1
    _borders(ws, tbl_s, ROW-1, 1, len(th))


# ══════════════════════════════════════════════════════════════════════
# PDF REPORT
# ══════════════════════════════════════════════════════════════════════

NAV  = colors.HexColor("#1F3864");  BLU  = colors.HexColor("#2E75B6")
RED  = colors.HexColor("#9C0006");  GRN  = colors.HexColor("#1E5631")
LGRN = colors.HexColor("#C6EFCE"); LRED = colors.HexColor("#FFC7CE")
LYLW = colors.HexColor("#FFEB9C"); WHT  = colors.white; BLK = colors.black


def _S():
    base = getSampleStyleSheet()
    return {
        "title":  ParagraphStyle("t",  parent=base["Normal"], fontSize=18, textColor=WHT,
                                 alignment=TA_CENTER, fontName="Helvetica-Bold", spaceAfter=2),
        "sub":    ParagraphStyle("s",  parent=base["Normal"], fontSize=10,
                                 textColor=colors.HexColor("#CCDDEE"), alignment=TA_CENTER),
        "sec":    ParagraphStyle("se", parent=base["Normal"], fontSize=12, textColor=WHT,
                                 fontName="Helvetica-Bold", alignment=TA_LEFT),
        "body":   ParagraphStyle("b",  parent=base["Normal"], fontSize=9, textColor=BLK,
                                 leading=13, spaceAfter=4),
        "small":  ParagraphStyle("sm", parent=base["Normal"], fontSize=8,
                                 textColor=colors.HexColor("#555555"), leading=11),
        "good":   ParagraphStyle("g",  parent=base["Normal"], fontSize=12, textColor=GRN,
                                 fontName="Helvetica-Bold", alignment=TA_CENTER, spaceBefore=8),
        "footer": ParagraphStyle("f",  parent=base["Normal"], fontSize=7,
                                 textColor=colors.HexColor("#888888"), alignment=TA_CENTER),
    }


def _sec_bar(title, W, S, bg=None):
    t = Table([[Paragraph(title, S["sec"])]], colWidths=[W])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), bg or NAV),
        ("TOPPADDING",    (0,0),(-1,-1), 7),
        ("BOTTOMPADDING", (0,0),(-1,-1), 7),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
    ]))
    return t


def _base_ts(hdr_bg=None):
    return TableStyle([
        ("BACKGROUND",     (0,0),(-1,0),  hdr_bg or NAV),
        ("TEXTCOLOR",      (0,0),(-1,0),  WHT),
        ("FONTNAME",       (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",       (0,0),(-1,-1), 8),
        ("ALIGN",          (0,0),(-1,-1), "CENTER"),
        ("ALIGN",          (0,1),(0,-1),  "LEFT"),
        ("FONTNAME",       (0,1),(-1,-1), "Helvetica"),
        ("ROWBACKGROUNDS", (0,1),(-1,-1), [WHT, colors.HexColor("#EBF3FB")]),
        ("GRID",           (0,0),(-1,-1), 0.4, colors.HexColor("#BFBFBF")),
        ("VALIGN",         (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",     (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",  (0,0),(-1,-1), 4),
        ("LEFTPADDING",    (0,0),(-1,-1), 5),
    ])


def build_pdf_report(stats, new_reviews, prev_date):
    S   = _S()
    doc = SimpleDocTemplate(
        OUTPUT_PDF, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=12*mm,  bottomMargin=15*mm,
    )
    W     = doc.width
    story = []

    # ── Header ──
    story.append(Table(
        [[Paragraph("iDEALZ — Weekly Google Reviews Report", S["title"])]],
        colWidths=[W],
        style=TableStyle([("BACKGROUND",(0,0),(-1,-1),NAV),
                          ("TOPPADDING",(0,0),(-1,-1),14),
                          ("BOTTOMPADDING",(0,0),(-1,-1),6)])
    ))
    story.append(Table(
        [[Paragraph(f"Week ending: {TODAY}   |   compared with: {prev_date}", S["sub"])]],
        colWidths=[W],
        style=TableStyle([("BACKGROUND",(0,0),(-1,-1),BLU),
                          ("TOPPADDING",(0,0),(-1,-1),5),
                          ("BOTTOMPADDING",(0,0),(-1,-1),5)])
    ))
    story.append(Spacer(1, 8*mm))

    # ── 1. Week at a Glance ──
    story.append(_sec_bar("1. THIS WEEK AT A GLANCE", W, S))
    story.append(Spacer(1, 2*mm))

    kpis = [
        ("Total New\nReviews",    stats["total"],         WHT, NAV),
        ("Critical\n(1-2 stars)", stats["critical"],      BLK, LRED if stats["critical"]      else LGRN),
        ("Needs Att.\n(3 stars)", stats["low"],           BLK, LYLW if stats["low"]           else LGRN),
        ("Positive\n(4-5 stars)", stats["good"],          BLK, LGRN),
        ("Negative\nSentiment",   stats["negative"],      BLK, LRED if stats["negative"]      else LGRN),
        ("Without\nOwner Reply",  stats["without_reply"], BLK, LRED if stats["without_reply"] else LGRN),
    ]
    kpi_row = []
    for label, value, tc, bg in kpis:
        kpi_row.append([
            Paragraph(f"<b>{label}</b>",
                      ParagraphStyle("kl", fontSize=8,
                                     textColor=colors.HexColor("#555555"),
                                     alignment=TA_CENTER)),
            Paragraph(f"<b>{value}</b>",
                      ParagraphStyle("kv", fontSize=18, textColor=tc,
                                     fontName="Helvetica-Bold", alignment=TA_CENTER)),
        ])
    kpi_tbl = Table([kpi_row], colWidths=[W/6]*6)
    kpi_ts  = [
        ("BOX",           (0,0),(-1,-1), 0.5, colors.HexColor("#BFBFBF")),
        ("INNERGRID",     (0,0),(-1,-1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 6),
        ("BOTTOMPADDING", (0,0),(-1,-1), 6),
    ]
    for i, (_, __, ___, bg) in enumerate(kpis):
        kpi_ts.append(("BACKGROUND", (i,0),(i,0), bg))
    kpi_tbl.setStyle(TableStyle(kpi_ts))
    story.append(kpi_tbl)
    story.append(Spacer(1, 6*mm))

    # ── 2. Per-Store Breakdown ──
    story.append(_sec_bar("2. PER STORE BREAKDOWN", W, S))
    story.append(Spacer(1, 2*mm))

    store_data = [["Branch", "New\nReviews", "Critical\n(1-2 stars)",
                   "Needs Att.\n(3 stars)", "Positive\n(4-5 stars)",
                   "Avg Rating", "Neg\nSentiment"]]
    for s in STORES:
        sname = s["name"]
        sd    = stats["by_store"].get(sname, {})
        tot   = sd.get("total", 0)
        crit  = sd.get("critical", 0)
        low_s = len([r for r in stats["critical_list"]
                     if r["store"] == sname and (r.get("rating") or 0) == 3])
        good  = tot - crit - low_s
        avg   = sd.get("avg")
        neg_s = sd.get("negative", 0)
        store_data.append([sname, str(tot), str(crit), str(low_s),
                           str(good), str(avg) if avg else "N/A", str(neg_s)])

    st_ts = list(_base_ts().getCommands())
    for ri in range(1, len(store_data)):
        if int(store_data[ri][2]) > 0:
            st_ts.append(("BACKGROUND", (2,ri),(2,ri), LRED))
    st_tbl = Table(store_data,
                   colWidths=[W*0.28, W*0.1, W*0.12, W*0.12, W*0.12, W*0.13, W*0.13])
    st_tbl.setStyle(TableStyle(st_ts))
    story.append(st_tbl)
    story.append(Spacer(1, 6*mm))

    # ── 3. Star Distribution ──
    story.append(_sec_bar("3. STAR RATING DISTRIBUTION  (This Week)", W, S))
    story.append(Spacer(1, 2*mm))

    total_new = stats["total"] or 1
    star_data = [["Stars", "Count", "% of Reviews", "Bar"]]
    star_clr  = {5:LGRN, 4:colors.HexColor("#DDEBF7"), 3:LYLW,
                 2:colors.HexColor("#FCE4D6"), 1:LRED}
    for star in [5, 4, 3, 2, 1]:
        cnt = stats["star_dist"].get(star, 0)
        pct = round(100 * cnt / total_new, 1)
        star_data.append([f"{star} stars", str(cnt), f"{pct}%",
                          chr(9608) * int(pct / 4)])
    star_ts = list(_base_ts().getCommands())
    for ri in range(1, 6):
        star_ts.append(("BACKGROUND", (0,ri),(-1,ri), star_clr[[5,4,3,2,1][ri-1]]))
    star_tbl = Table(star_data, colWidths=[W*0.15, W*0.12, W*0.2, W*0.53])
    star_tbl.setStyle(TableStyle(star_ts))
    story.append(star_tbl)

    # ── 4. Critical Reviews ──
    story.append(PageBreak())
    critical_list = stats["critical_list"]
    story.append(_sec_bar(
        f"4. CRITICAL REVIEWS  (1-2 Stars)  —  {len(critical_list)} this week",
        W, S, bg=RED if critical_list else GRN
    ))
    story.append(Spacer(1, 3*mm))

    if not critical_list:
        story.append(Paragraph("No critical reviews (1-2 stars) this week!", S["good"]))
    else:
        for r in critical_list:
            rat     = r.get("rating") or 0
            replied = r.get("owner_reply","").strip()
            bg_hex  = "#FFC7CE" if rat <= 2 else "#FFF2CC"
            h_tbl   = Table([[Paragraph(
                f"<b>{r.get('store','')}  |  {rat} stars  |  "
                f"{r.get('author','')}  |  {r.get('date_raw','')}</b>",
                ParagraphStyle("ch", fontSize=9, textColor=BLK,
                               fontName="Helvetica-Bold")
            )]], colWidths=[W])
            h_tbl.setStyle(TableStyle([
                ("BACKGROUND",    (0,0),(-1,-1), colors.HexColor(bg_hex)),
                ("TOPPADDING",    (0,0),(-1,-1), 4),
                ("BOTTOMPADDING", (0,0),(-1,-1), 4),
                ("LEFTPADDING",   (0,0),(-1,-1), 6),
                ("BOX",           (0,0),(-1,-1), 0.5, colors.HexColor("#AAAAAA")),
            ]))
            story.append(h_tbl)
            story.append(Paragraph(r.get("text","No review text."), S["body"]))
            if not replied:
                story.append(Paragraph(
                    "<b><font color='#9C0006'>ACTION REQUIRED: No owner reply yet</font></b>",
                    S["small"]
                ))
            else:
                story.append(Paragraph(
                    f"<i>Owner reply: {replied[:250]}{'...' if len(replied)>250 else ''}</i>",
                    S["small"]
                ))
            story.append(Spacer(1, 3*mm))

    # ── 5. All New Reviews ──
    if new_reviews:
        story.append(PageBreak())
        story.append(_sec_bar(f"5. ALL NEW REVIEWS THIS WEEK  ({len(new_reviews)})", W, S))
        story.append(Spacer(1, 2*mm))

        tbl_data = [["Store", "Stars", "Sentiment", "Review Text", "Date", "Replied?"]]
        for r in sorted(new_reviews,
                        key=lambda r: (r.get("sentiment","") != "Negative",
                                       -(r.get("rating") or 3))):
            tbl_data.append([
                r.get("store",""),
                str(r.get("rating","")),
                r.get("sentiment",""),
                (r.get("text","") or "")[:180],
                r.get("date_raw",""),
                "Yes" if r.get("owner_reply","").strip() else "No",
            ])
        nr_ts = list(_base_ts().getCommands())
        for ri, r in enumerate(new_reviews, 1):
            sent = r.get("sentiment","")
            clr  = LRED if sent == "Negative" else (LGRN if sent == "Positive" else LYLW)
            nr_ts.append(("BACKGROUND", (2,ri),(2,ri), clr))
        nr_tbl = Table(tbl_data,
                       colWidths=[W*0.18, W*0.07, W*0.1, W*0.42, W*0.13, W*0.1],
                       repeatRows=1)
        nr_tbl.setStyle(TableStyle(nr_ts))
        story.append(nr_tbl)

    # ── 6. Topic Categories ──
    if stats["categories"]:
        story.append(Spacer(1, 6*mm))
        story.append(_sec_bar("6. TOPIC CATEGORIES MENTIONED", W, S))
        story.append(Spacer(1, 2*mm))
        cat_data = [["Category", "Count", "% of New Reviews"]]
        for cat, cnt in stats["categories"].most_common():
            pct = round(100 * cnt / (stats["total"] or 1), 1)
            cat_data.append([cat, str(cnt), f"{pct}%"])
        cat_tbl = Table(cat_data, colWidths=[W*0.55, W*0.2, W*0.25])
        cat_tbl.setStyle(_base_ts())
        story.append(cat_tbl)

    # ── Footer ──
    story.append(Spacer(1, 10*mm))
    story.append(HRFlowable(width=W, thickness=0.5, color=colors.HexColor("#AAAAAA")))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        f"Generated by iDealz Weekly Review Tracker  |  {TODAY}  "
        f"|  {len(new_reviews)} new reviews across {len(STORES)} branches",
        S["footer"]
    ))

    doc.build(story)
    print(f"  [OK] PDF saved -> {OUTPUT_PDF}")


# ══════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════

def main():
    print(f"\n{'='*60}")
    print(f"  IDEALZ WEEKLY TRACKER — {TODAY}")
    print(f"{'='*60}")

    # Step 1 — Scrape
    print("\n[1/4] Scraping latest reviews from Google Maps...")
    current = scrape_all_stores()
    if not current:
        print("  [X] No reviews scraped. Check store URLs.")
        return
    print(f"  Total scraped: {len(current)}")

    # Step 2 — Save snapshot
    print("\n[2/4] Saving snapshot...")
    save_snapshot(current)

    # Step 3 — Load previous snapshot
    print("\n[3/4] Loading previous snapshot...")
    previous, prev_date = load_previous_snapshot()
    if not previous:
        print("  [!] No previous snapshot found.")
        print("  [!] This is your BASELINE run — no report generated yet.")
        print("  [!] Run again next Monday to get your first weekly report.")
        return

    # Step 4 — Analyse + generate reports
    print("\n[4/4] Analysing and generating reports...")
    enrich(current)
    enrich(previous)

    new_reviews = get_new_reviews(current, previous)
    enrich(new_reviews)
    stats = weekly_stats(new_reviews)

    # Print summary to console
    print(f"\n  ── WEEK SUMMARY ──────────────────────────────────────")
    print(f"  New reviews this week  : {stats['total']}")
    print(f"  Critical (1-2 stars)   : {stats['critical']}")
    print(f"  Needs attention (3 str): {stats['low']}")
    print(f"  Positive (4-5 stars)   : {stats['good']}")
    print(f"  Negative sentiment     : {stats['negative']}")
    print(f"  Without owner reply    : {stats['without_reply']}")
    print(f"  ─────────────────────────────────────────────────────")
    for s in STORES:
        sd = stats["by_store"].get(s["name"], {})
        print(f"  {s['name']}")
        print(f"    New: {sd.get('total',0)}  |  "
              f"Critical: {sd.get('critical',0)}  |  "
              f"Avg: {sd.get('avg','N/A')} stars")
    print(f"  ─────────────────────────────────────────────────────\n")

    all_snapshots = load_all_snapshots()
    build_excel_report(stats, new_reviews, prev_date, all_snapshots)
    build_pdf_report(stats, new_reviews, prev_date)

    print(f"\n{'='*60}")
    print(f"  DONE — Files saved in this folder:")
    print(f"  PDF:   {OUTPUT_PDF}")
    print(f"  Excel: {OUTPUT_XLSX}")
    print(f"  Snap:  snapshots/reviews_{TODAY}.json")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()